import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import string  
import os
from prettytable import PrettyTable
from scipy.stats import skew, kurtosis

wb = load_workbook("C:/Users/User/Desktop/static/homework1_descriptive_statistics_and_boxplot/static_data2.xlsx")
ws = wb.active

companies = ["國泰金控", "台積電"]
variables = ["單月營收(千元)", "每股盈餘(元)", "員工人數(人)", "資產總額(元)", "實收資本額(元)"]

plt.rcParams['font.family'] = 'MingLiU'


# Function to convert column letter to index and vice versa
def column_letter_to_index(letter):
    return string.ascii_uppercase.index(letter) + 1
#string.ascii_uppercase.index(letter) : convert uppercase letter to index but A to 0.
# So we need to add 1 to get the correct index 
def index_to_column_letter(index):
    return string.ascii_uppercase[index - 1]
#string.ascii_uppercase: A to 1 
#but list index start from 0.

def extract_data(start_col, row_count=50):
    data = {var: [] for var in variables} #dictionary {'單月營收(千元)': [], '每股盈餘': [],...}
    start_col_index = column_letter_to_index(start_col)  # Get the starting column index
    for row in range(2, row_count):  # Skip header row
        for i, var in enumerate(variables):
            value = ws[f"{index_to_column_letter(start_col_index + i)}{row}"].value
            if value is not None:
                data[var].append(value)
    return data

# Extract data for both companies
data = {}
start_columns = {'國泰金控': 'C', '台積電': 'K'} # key : value 
for company in companies:
    data[company] = extract_data(start_columns[company])


output_dir = "C:/Users/User/Desktop/static/homework1_descriptive_statistics_and_boxplot/output2"  
os.makedirs(output_dir, exist_ok=True)  
stats_file = os.path.join(output_dir, "descriptive_statistics_new.txt")

for var in variables:
    plt.figure(figsize=(8, 6))
    boxprops = dict(facecolor='lightblue', color='blue')  # box color
    medianprops = dict(color='red')  # median color
    flierprops = dict(marker='x', color='green', markersize=5)  
    capprops = dict(color='black')  
    whiskerprops = dict(color='black')  
    
    boxplot_result = plt.boxplot(
        [data[company][var] for company in companies],
        labels=companies,
        boxprops=boxprops,
        medianprops=medianprops,
        flierprops=flierprops,
        capprops=capprops,
        whiskerprops=whiskerprops,
        patch_artist=True
    )
    # plt.boxplot(
    #     [data[company][var] for company in companies],
    #     labels=companies,
    #     boxprops=boxprops,
    #     medianprops=medianprops,
    #     flierprops=flierprops,
    #     capprops=capprops,
    #     whiskerprops=whiskerprops,
    #     patch_artist=True  
    # )
    
    plt.ylabel(var)
    # plt.title(f"{var} Comparison")
    plt.grid()
    for i, company in enumerate(companies):
        outliers = boxplot_result['fliers'][i].get_ydata()  # 獲取離群值的數據
        if len(outliers) > 0:
            for outlier in outliers:
                plt.annotate(f'{outlier:.2f}', xy=(i + 1, outlier), xytext=(5, 5),
                             textcoords='offset points', color='red')  # 標註離群值
    # Save the figure
    plt.savefig(os.path.join(output_dir, f"{var}.png"), dpi=300, bbox_inches='tight')
    plt.close()  # Close the figure to free up memory
'''
with open(stats_file, "w", encoding='utf-8') as f:
    for var in variables:
        table = PrettyTable()
        table.field_names = ["Company", "Mean", "Median", "Max","Min",
                             "range","Std Dev", "IQR", "Skewness"]
                            # "Lower inner fence", "Upper inner fence", 
                            # "Lower outer fence", "Upper outer fence",
        
        for company in companies:
            values = np.array(data[company][var])
            
            # Descriptive statistics calculations
            mean = np.mean(values)
            median = np.median(values)
            std_dev = np.std(values)
            q1 = np.percentile(values, 25)
            q3 = np.percentile(values, 75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            second_lower_bound = q1 - 3 * iqr
            second_upper_bound = q3 + 3 * iqr

            table.add_row([
                company,
                round(mean, 2),
                round(median, 2),
                round(max(values), 2),
                round(min(values), 2),
                round(max(values) - min(values), 2),
                round(std_dev, 2),
                # round(q1, 2),
                # round(q3, 2),
                round(iqr, 2),
                round(skew(values), 2)
                # round(lower_bound, 2),
                # round(upper_bound, 2),
                # round(second_lower_bound, 2),
                # round(second_upper_bound, 2),
                
                
                
            ])
        
        # Print and save the statistics for the current variable
        f.write(f"\nDescriptive Statistics for {var}:\n")
        f.write(str(table))
        f.write("\n\n")
        # print(f"Descriptive Statistics for {var}:\n")
        # print(table)
'''